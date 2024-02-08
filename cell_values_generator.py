import random
import openpyxl as xl
from openpyxl.chart import Reference, BarChart


def generate_cell_values(rows_required):
    wb = xl.load_workbook('transactions.xlsx')
    sheet = wb['Sheet1']
    row_count = sheet.max_row
    col_count = sheet.max_column

    for row in range(row_count, rows_required):
        col = 1
        cell = sheet.cell(row, col)
        cell.value = sheet.cell(row - 1, col).value + 1

        col = 2
        cell = sheet.cell(row, col)
        cell.value = random.randint(1,4)

        col = 3
        cell = sheet.cell(row, col)
        cell.value = random.randint(10, 100)
    wb.save('transactions_generated.xlsx')


def calc_discounted_price():
    wb = xl.load_workbook('transactions_generated.xlsx')
    sheet = wb['Sheet1']
    row_count = sheet.max_row
    col_count = sheet.max_column
    sheet.cell(1, 4).value = 'discounted_price'

    for row in range(2, row_count + 1):
        price = sheet.cell(row, 3).value
        corrected_price = price * .9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price
    wb.save('transactions_generated.xlsx')


def generate_bar_chart():
    wb = xl.load_workbook('transactions_generated.xlsx')
    sheet = wb['Sheet1']
    sheet_2 = wb.create_sheet('Bar_Chart')
    row_count = sheet.max_row
    col_count = sheet.max_column
    values = Reference(sheet,
                       min_row=2,
                       max_row=row_count,
                       min_col=4,
                       max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet_2.add_chart(chart, 'a2')
    wb.save('transactions_generated.xlsx')

def generate_total_price():
    wb = xl.load_workbook('transactions_generated.xlsx')
    sheet = wb['Sheet1']
    row_count = sheet.max_row
    col_count = sheet.max_column
    total_price_cell = sheet.cell(row_count+1, 4)
    total_price = 0
    # for row in range(2, row_count + 1):
    #     price = sheet.cell(row, 4).value
    #     total_price += price
    total_price_formula = f'=SUM(D2: D{row_count})'
    total_price_cell.value = total_price_formula
    wb.save('transactions_generated.xlsx')